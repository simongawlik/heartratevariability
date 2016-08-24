
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


from bs4 import BeautifulSoup
from openpyxl import load_workbook
import settings
import math
import os


data_path = os.path.join(os.getcwd(), "workbooks")

# # To prevent download dialog
# profile = webdriver.FirefoxProfile()
# profile.set_preference('browser.download.folderList', 2)
# profile.set_preference('browser.download.manager.showWhenStarting', False)
# profile.set_preference('browser.download.dir', data_path)
# profile.set_preference('browser.helperApps.neverAsk.saveToDisk', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
# driver = webdriver.Firefox(profile)

# url = "http://www.movescount.com/auth?redirect_uri=%2foverview"

# driver.implicitly_wait(10) # seconds
# driver.get(url)
# driver.find_element_by_name("email").send_keys(settings.email)
# driver.find_element_by_name("password").send_keys(settings.password)
# driver.find_element_by_name("password").send_keys(Keys.RETURN)

# try:
# 	link = WebDriverWait(driver, 10).until(
# 		EC.presence_of_element_located((By.CLASS_NAME, "h3"))
# 	)
# finally:
# 	print("waited for workouts")	

# workouts = driver.find_elements_by_css_selector('a.h3')

# workouts[0].click()

# try:
# 	link = WebDriverWait(driver, 10).until(
# 		EC.presence_of_element_located((By.XPATH, "//a[@data-export-format='xlsx']"))
# 	)
# finally:
# 	print("waited for download button")

# menu = driver.find_elements_by_xpath("//a[@class='link link--light middle-all']")
# link_to_download = driver.find_elements_by_xpath("//a[@data-export-format='xlsx']")

# actions = webdriver.ActionChains(driver)
# actions.move_to_element(menu[0])
# actions.click(link_to_download[0])
# actions.perform()

# driver.close()

def get_new_most_recent_file_index(most_recent, most_recent_index, next, next_index):
	for i in range(1, 6):
		print(current_file_date[i])
		print(most_recent_file_date[i])
		if current_file_date[i] > most_recent_file_date[i]:
			return next_index
		elif current_file_date[i] < most_recent_file_date[i]:
			return most_recent_index
	return most_recent_index


list_of_files = os.listdir(data_path)

filtered_list = []

for file_name in list_of_files:
	if file_name[0] == 'M':
		filtered_list.append(file_name)


print(filtered_list)

most_recent_file_date = filtered_list[0].split("_")
most_recent_file_index = 0

for i in range(1, len(filtered_list)):
	current_file_date = filtered_list[i].split("_")
	most_recent_file_index = get_new_most_recent_file_index(
		most_recent_file_date, most_recent_file_index,
		current_file_date, i)
	print(most_recent_file_index)
	most_recent_file_date = filtered_list[most_recent_file_index].split("_")
	

print(filtered_list[most_recent_file_index])


file_name = filtered_list[most_recent_file_index]

wb = load_workbook(os.path.join(data_path, file_name))

x_values = []
x_sum = 0

ws = wb.active

start_row = 3
start_column = 40
measurement_length = 180


for row in range(start_row, start_row + measurement_length):
	next_cell = ws.cell(column=start_column, row=row).value
	#print(x_values)
	x_values.append(next_cell)
	x_sum = x_sum + next_cell


mean = x_sum / measurement_length

sum_of_squared_errors = 0
for x in x_values:
	sq_error = x - mean
	sum_of_squared_errors = sum_of_squared_errors + (sq_error * sq_error)


variance = sum_of_squared_errors / (measurement_length - 1)
std_deviation = math.sqrt(variance)

print(std_deviation)








